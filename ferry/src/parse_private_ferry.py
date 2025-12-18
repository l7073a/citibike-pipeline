"""
Private Ferry Excel Parser - Robust Multi-Era Version

Handles structural variations across 2013-2025:
- 2013-2019: Headers in rows 0+2, data starts row 4+
- 2024+: Headers in rows 0+1, data starts row 2

Key features:
1. Auto-detects header rows by finding "Day"/"Date" pattern
2. Auto-detects data start by finding first date value
3. Skips "Total" columns
4. Handles merged cells via forward-fill
5. Extracts holiday annotations from day names
6. Validates against Monthly Totals sheet

Usage:
    python parse_private_ferry.py <excel_file> [--output <parquet_file>]

    # Process all files:
    python parse_private_ferry.py --all --output-dir ferry/data/private_ferry/parsed/
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import re
from openpyxl import load_workbook
import sys
import argparse
from typing import Optional, Tuple, List, Dict, Any


def extract_holiday(day_str: str) -> Tuple[str, Optional[str]]:
    """
    Extract holiday name from strings like "(MLK Day) Monday".
    Returns: (clean_day_name, holiday_name or None)
    """
    if not isinstance(day_str, str):
        return (str(day_str).strip() if day_str else '', None)

    day_str = day_str.strip()

    # Match patterns like "(New Year's Day) Monday"
    match = re.match(r'\(([^)]+)\)\s*(\w+)', day_str)
    if match:
        return (match.group(2).strip(), match.group(1).strip())

    return (day_str, None)


def is_date_value(val) -> bool:
    """Check if a value looks like a date."""
    if isinstance(val, datetime):
        return True
    if val is None:
        return False
    try:
        pd.to_datetime(val)
        return True
    except:
        return False


def find_header_and_data_rows(ws) -> Tuple[int, int, int]:
    """
    Find the header rows and data start row by scanning the sheet.

    Returns: (route_row, stop_row, data_start_row)

    Handles two structure types:
    1. Modern (2024+): Row 0 = title + routes, Row 1 = Day/Date + stops, Row 2+ = data
    2. Legacy (2013-2019): Row 0 = Day/Date + routes, Row 2 = stops, Row 4+ = data
    """
    rows_data = list(ws.iter_rows(max_row=12, values_only=True))

    day_date_row = None
    route_row = None
    stop_row = None

    # Step 1: Find the row with "Day" and/or "Date" in columns 0-1
    for i, row in enumerate(rows_data):
        col0 = str(row[0]).strip().lower() if row[0] else ''
        col1 = str(row[1]).strip().lower() if len(row) > 1 and row[1] else ''

        if col0 == 'day' or col1 == 'date':
            day_date_row = i
            break

    if day_date_row is None:
        # Fallback
        day_date_row = 0

    # Step 2: Determine structure by checking the row AFTER day_date_row
    # - Modern (2024+): Next row is DATA (has date in col 1)
    # - Legacy (2013-2019): Next row is empty, then stops, then data
    next_row_is_data = False
    if day_date_row + 1 < len(rows_data):
        next_row = rows_data[day_date_row + 1]
        next_row_is_data = len(next_row) > 1 and is_date_value(next_row[1])

    if next_row_is_data:
        # Modern structure: Day/Date row has stops, previous row has routes
        # Row N-1: routes, Row N: Day/Date + stops, Row N+1: data
        stop_row = day_date_row
        for i in range(day_date_row - 1, -1, -1):
            row = rows_data[i]
            has_routes = any(
                val is not None and str(val).strip() and str(val).strip().lower() not in ['none', 'weekday', 'weekdays', '']
                for val in row[2:15]
            )
            if has_routes:
                route_row = i
                break
        if route_row is None:
            route_row = max(0, day_date_row - 1)
    else:
        # Legacy structure: Day/Date row has ORIGINS, destinations are ABOVE
        # Row 0: Destinations (Pier 79, WFC, etc.)
        # Row 2: Day/Date + Origins (Paulus Hook, etc.)
        # Row 3+: Empty or data
        stop_row = day_date_row  # Origins are in the Day/Date row

        # Look BACKWARD for destination row (row with content before Day/Date row)
        for i in range(day_date_row - 1, -1, -1):
            row = rows_data[i]
            has_dests = any(
                val is not None and str(val).strip() and str(val).strip().lower() not in ['none', '']
                for val in row[2:15]
            )
            if has_dests:
                route_row = i
                break
        if route_row is None:
            route_row = 0  # Fallback to row 0

    # Step 3: Find data start (first row after stop_row where col 1 is a date)
    # Legacy files can have many empty rows between headers and data
    data_start = stop_row + 1
    for i in range(stop_row + 1, min(len(rows_data), stop_row + 10)):
        row = rows_data[i]
        if len(row) > 1 and is_date_value(row[1]):
            data_start = i
            break

    return (route_row, stop_row, data_start)


def forward_fill_header(header_row: list, skip_patterns: List[str] = None) -> list:
    """
    Forward-fill None values in a header row (handles merged cells).

    Args:
        header_row: List of header values
        skip_patterns: Patterns to treat as None (e.g., ['Weekday', 'Total'])
    """
    if skip_patterns is None:
        skip_patterns = ['weekday', 'weekdays', 'total', 'none', 'nan', '']

    result = []
    last_value = None

    for val in header_row:
        val_str = str(val).strip().lower() if val is not None else ''

        if val is not None and val_str not in skip_patterns:
            last_value = str(val).strip()

        result.append(last_value)

    return result


def parse_operator_sheet(wb, sheet_name: str, operator_name: str = None) -> pd.DataFrame:
    """
    Parse a single operator sheet with auto-detected headers.

    Returns DataFrame with columns:
        date, day_of_week, holiday, operator, destination, origin, ridership
    """
    ws = wb[sheet_name]

    # Find header structure
    route_row_idx, stop_row_idx, data_start_idx = find_header_and_data_rows(ws)

    # Read header rows
    all_rows = list(ws.iter_rows(values_only=True))

    if route_row_idx >= len(all_rows) or stop_row_idx >= len(all_rows):
        print(f"    Warning: Could not find headers in {sheet_name}")
        return pd.DataFrame()

    route_row = list(all_rows[route_row_idx])
    stop_row = list(all_rows[stop_row_idx])

    # Forward-fill routes (handles merged cells)
    routes_filled = forward_fill_header(route_row)

    # Clean stop names
    stops = []
    for val in stop_row:
        if val is None:
            stops.append(None)
        else:
            s = str(val).strip()
            if s.lower() in ['day', 'date', 'weekday', 'weekdays', 'none', '']:
                stops.append(None)
            else:
                stops.append(s)

    # Build column metadata (skip first 2 columns: Day, Date)
    columns_meta = []
    for i in range(2, min(len(routes_filled), len(stops))):
        route = routes_filled[i]
        stop = stops[i]

        # Skip if stop is None - these are spacer/total columns
        # A valid data column MUST have a stop name
        if stop is None:
            continue

        # Skip if both are None (redundant now but kept for clarity)
        if route is None and stop is None:
            continue

        # Skip "Total" columns
        if route and 'total' in route.lower():
            continue
        if stop and 'total' in stop.lower():
            continue

        columns_meta.append({
            'col_idx': i,
            'destination': route or 'Unknown',  # Route/destination (Manhattan pier)
            'origin': stop                       # Origin/stop (NJ departure or boarding stop)
        })

    # Parse data rows
    records = []

    for row_idx in range(data_start_idx, len(all_rows)):
        row = all_rows[row_idx]

        if not row or len(row) < 2:
            continue

        day_val = row[0]
        date_val = row[1]

        # Skip empty rows
        if day_val is None and date_val is None:
            continue

        # Skip summary rows
        if day_val and isinstance(day_val, str):
            day_lower = day_val.lower()
            if any(skip in day_lower for skip in ['total', 'average', 'weekday ridership', 'number of']):
                continue

        # Parse date
        if isinstance(date_val, datetime):
            date_parsed = date_val.date()
        elif date_val:
            try:
                date_parsed = pd.to_datetime(date_val).date()
            except:
                continue  # Skip rows without valid dates
        else:
            continue

        # Extract holiday
        clean_day, holiday = extract_holiday(str(day_val) if day_val else '')

        # Extract ridership for each column
        for col_meta in columns_meta:
            col_idx = col_meta['col_idx']
            if col_idx >= len(row):
                continue

            value = row[col_idx]

            # Convert to integer ridership
            if value is None or value == '' or (isinstance(value, str) and value.strip() == ''):
                ridership = 0
            elif isinstance(value, (int, float)):
                ridership = int(value) if not pd.isna(value) else 0
            else:
                try:
                    ridership = int(float(str(value).replace(',', '')))
                except:
                    ridership = 0

            records.append({
                'date': date_parsed,
                'day_of_week': clean_day,
                'holiday': holiday,
                'operator': operator_name or sheet_name,
                'destination': col_meta['destination'],
                'origin': col_meta['origin'],
                'ridership': ridership
            })

    return pd.DataFrame(records)


def parse_monthly_totals(wb) -> Dict[str, Any]:
    """Parse the Monthly Totals sheet for validation."""
    if 'Monthly Totals' not in wb.sheetnames:
        return {}

    ws = wb['Monthly Totals']

    summary = {'by_operator': {}}
    section = None

    for row in ws.iter_rows(values_only=True):
        if not row or not any(row):
            continue

        # Find the cell with content (could be col 0, 1, or 2)
        name = None
        value = None

        for i, cell in enumerate(row[:4]):
            if cell is None:
                continue
            cell_str = str(cell).strip()

            if 'Ridership by Operator' in cell_str:
                section = 'operator'
                break
            elif 'Ridership by' in cell_str:
                section = None  # End of operator section
                break
            elif section == 'operator' and cell_str and cell_str != 'Total':
                if isinstance(cell, (int, float)) and not pd.isna(cell):
                    value = cell
                elif name is None:
                    name = cell_str

        if section == 'operator' and name and value:
            summary['by_operator'][name] = int(value)

    return summary


# Operator name mappings (normalize across years)
OPERATOR_MAPPINGS = {
    # 2024 names
    'NYC Ferry': 'NYC Ferry',
    'NYWW (Port Imperial FC)': 'NY Waterway',
    'New York Water Taxi': 'Water Taxi',
    'SeaStreak': 'SeaStreak',
    'Liberty Landing Ferry': 'Liberty Landing',
    # Older names
    'NY Waterway': 'NY Waterway',
    'NY Waterway-(Port Imperial FC)': 'NY Waterway',
    'NY Waterway-(Billy Bey FC)': 'Billy Bey',
    'Billy Bey': 'Billy Bey',
    'Water Tours': 'Water Tours',
    'New York Water Tours': 'Water Tours',
    'Baseball': 'Baseball',
    'HMS': 'HMS',
}

# Sheets to skip
SKIP_SHEETS = {'Monthly Totals', 'Weekday Totals', 'Sheet1', 'Sheet2'}


def parse_private_ferry_excel(filepath, verbose: bool = True) -> Dict[str, Any]:
    """
    Parse a complete private ferry ridership Excel file.

    Returns dict with:
        - 'data': Combined DataFrame (long format)
        - 'summary': Monthly totals for validation
        - 'operators': List of operators found
        - 'file_info': Metadata
    """
    filepath = Path(filepath)
    wb = load_workbook(filepath, data_only=True)

    all_data = []
    operators_found = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        operator = OPERATOR_MAPPINGS.get(sheet_name, sheet_name)
        operators_found.append(operator)

        if verbose:
            print(f"  Parsing: {sheet_name} -> {operator}")

        try:
            df = parse_operator_sheet(wb, sheet_name, operator)
            if len(df) > 0:
                all_data.append(df)
                if verbose:
                    print(f"    -> {len(df):,} records, {df['ridership'].sum():,} total ridership")
            else:
                if verbose:
                    print(f"    -> No data found")
        except Exception as e:
            if verbose:
                print(f"    -> ERROR: {e}")

    # Combine all data
    combined = pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()

    # Parse summary
    try:
        summary = parse_monthly_totals(wb)
    except:
        summary = {}

    # Extract file metadata
    match = re.search(r'(\d{4})_(\d{2})', filepath.stem)
    file_year = int(match.group(1)) if match else None
    file_month = int(match.group(2)) if match else None

    return {
        'data': combined,
        'summary': summary,
        'operators': operators_found,
        'date_range': (combined['date'].min(), combined['date'].max()) if len(combined) > 0 else (None, None),
        'file_info': {
            'filename': filepath.name,
            'year': file_year,
            'month': file_month,
            'total_records': len(combined),
            'total_ridership': int(combined['ridership'].sum()) if len(combined) > 0 else 0
        }
    }


def validate_against_summary(result: Dict[str, Any]) -> Dict[str, Any]:
    """Validate parsed data against Monthly Totals."""
    if not result['summary'] or result['data'].empty:
        return {'status': 'skipped'}

    data = result['data']
    summary = result['summary']

    validation = {'status': 'passed', 'checks': [], 'warnings': []}

    # Aggregate by operator
    parsed_totals = data.groupby('operator')['ridership'].sum().to_dict()

    for summary_op, expected in summary.get('by_operator', {}).items():
        # Map summary operator name to our normalized name
        op_normalized = OPERATOR_MAPPINGS.get(summary_op.strip(), summary_op.strip())
        actual = parsed_totals.get(op_normalized, 0)

        if expected == 0:
            continue

        diff_pct = abs(actual - expected) / expected * 100

        check = {
            'operator': summary_op,
            'expected': expected,
            'actual': actual,
            'diff_pct': round(diff_pct, 1)
        }
        validation['checks'].append(check)

        if diff_pct > 10:
            validation['status'] = 'warning'
            validation['warnings'].append(f"{summary_op}: {diff_pct:.1f}% diff (expected {expected:,}, got {actual:,})")

    return validation


def process_all_files(input_dir: Path, output_dir: Path, verbose: bool = True) -> pd.DataFrame:
    """Process all Excel files in a directory."""
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    all_results = []
    errors = []

    files = sorted(input_dir.glob("*Ridership*.xlsx"))
    files = [f for f in files if 'summary' not in f.name.lower() and 'CY' not in f.name and 'CalendarYear' not in f.name]

    print(f"Processing {len(files)} files...")

    for filepath in files:
        print(f"\n{'='*60}")
        print(f"File: {filepath.name}")

        try:
            result = parse_private_ferry_excel(filepath, verbose=verbose)

            if len(result['data']) > 0:
                all_results.append(result['data'])

                # Validate
                validation = validate_against_summary(result)
                if validation['status'] == 'warning':
                    print(f"  VALIDATION WARNINGS:")
                    for w in validation['warnings']:
                        print(f"    - {w}")

                print(f"  -> {result['file_info']['total_records']:,} records, {result['file_info']['total_ridership']:,} ridership")
            else:
                print(f"  -> No data extracted")
                errors.append((filepath.name, "No data"))

        except Exception as e:
            print(f"  -> ERROR: {e}")
            errors.append((filepath.name, str(e)))

    # Combine all results
    if all_results:
        combined = pd.concat(all_results, ignore_index=True)

        # Save combined output
        output_path = output_dir / "private_ferry_daily.parquet"
        combined.to_parquet(output_path, index=False)
        print(f"\n{'='*60}")
        print(f"SAVED: {output_path}")
        print(f"Total records: {len(combined):,}")
        print(f"Total ridership: {combined['ridership'].sum():,}")
        print(f"Date range: {combined['date'].min()} to {combined['date'].max()}")

        if errors:
            print(f"\nErrors ({len(errors)}):")
            for fname, err in errors:
                print(f"  - {fname}: {err}")

        return combined

    return pd.DataFrame()


def main():
    parser = argparse.ArgumentParser(description='Parse private ferry Excel files')
    parser.add_argument('input', nargs='?', help='Excel file or --all for batch processing')
    parser.add_argument('--all', action='store_true', help='Process all files in raw directory')
    parser.add_argument('--output', '-o', help='Output parquet file')
    parser.add_argument('--output-dir', help='Output directory for batch processing')
    parser.add_argument('--quiet', '-q', action='store_true', help='Minimal output')

    args = parser.parse_args()

    if args.all:
        input_dir = Path("ferry/data/private_ferry/raw")
        output_dir = Path(args.output_dir or "ferry/data/private_ferry")
        process_all_files(input_dir, output_dir, verbose=not args.quiet)
    elif args.input:
        result = parse_private_ferry_excel(args.input, verbose=not args.quiet)

        print(f"\n{'='*60}")
        print("RESULTS")
        print('='*60)
        print(f"Records: {result['file_info']['total_records']:,}")
        print(f"Ridership: {result['file_info']['total_ridership']:,}")
        print(f"Date range: {result['date_range']}")
        print(f"Operators: {', '.join(result['operators'])}")

        # Show sample
        print(f"\nSample data:")
        print(result['data'].head(10).to_string())

        # Ridership by operator
        print(f"\nBy operator:")
        print(result['data'].groupby('operator')['ridership'].sum().sort_values(ascending=False).to_string())

        # Save if requested
        if args.output:
            result['data'].to_parquet(args.output, index=False)
            print(f"\nSaved to: {args.output}")
    else:
        parser.print_help()


if __name__ == '__main__':
    main()
